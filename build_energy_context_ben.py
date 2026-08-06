"""Load national-level BEN (Balanço Energético Nacional, EPE) context tables
for the 4 priority chains into the Published layer.

Investigated and confirmed (2026-08-06): the BEN does not publish "Anexo IX"
or "Anexo X" in a pre-tidied ("tidyverse", one column for year + one for
value) format -- both are matrix workbooks (one sheet per year). The only
place EPE offers a tidy export is the separate BEN Interativo tool
(dashboard.epe.gov.br/apps/ben/), whose downloads are Shiny session-scoped
URLs, not stable enough to script against. So this extractor reads the
static "Capítulo" matrix workbooks instead (stable URLs, same underlying
numbers) and does the wide-to-long unpivot itself.

The BEN is a national/annual/sectoral reference (EPE's own accounting, not
survey microdata) -- it has no municipal or state granularity for these
sectors. It must NOT be cross-referenced row-by-row with
mv_published_indicators (which is municipal/firm-level via NCM/CNAE). This
mirrors the OBEPE precedent in the sibling Atlas Solar Justo project: a
national/state reference cited alongside municipal data, never merged into
it. The `analytical_energy_context_national` table this script loads is
intentionally keyed by (ano_referencia, cadeia_prioritaria, setor_ben,
fonte_energetica) -- no NCM/CNAE/municipal key -- so it cannot accidentally
be joined into the municipal Published tables.

Chain -> BEN table mapping (confirmed by inspecting both workbooks; exact
sector-level match, not an approximation):
  - aco                    -> Capítulo 3, Tabela 3.7.2.a
                               "SETOR INDUSTRIAL - FERRO-GUSA E AÇO" (mil tep)
  - fertilizantes           -> Capítulo 3, Tabela 3.7.5.a
                               "SETOR INDUSTRIAL - QUÍMICA" (mil tep)
  - silicio                 -> Capítulo 2, Tabela 2.8
                               "ENERGIA SOLAR FOTOVOLTAICA" (GWh)
  - combustiveis_transicao  -> Capítulo 2, Tabela 2.21 "BIODIESEL" (mil m³)
                               + Tabela 2.32 "ÁLCOOL ETÍLICO TOTAL" (mil m³)
                               (whole-chain theme, not one sub-sector -- see
                               project memory on the sector->BEN mapping)

Each ".a"/absolute block in Capítulo 3 has a paired ".b"/percentage block
right after it with the identical title; TARGET_BLOCKS below always matches
the first (".a", absolute) occurrence, which is what MATCH_OCCURRENCE=0
means. The workbooks self-document their unit in the title row (a cell
reading "UNIDADE: ...": "mil tep", "%", "GWh", "mil m³") -- read from there
instead of hard-coding it, so a wrong guess can't silently ship.

Usage:
    python build_energy_context_ben.py

Output:
    outputs/analytical_energy_context_ben/load_analytical_energy_context_national.sql
    (CREATE TABLE IF NOT EXISTS safety net + full DELETE/INSERT of this
    table's only owner, ready to run with `psql -f`)
"""

from __future__ import annotations

import json
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent
INPUTS_DIR = ROOT / "inputs" / "official"
OUTPUT_DIR = ROOT / "outputs" / "analytical_energy_context_ben"

CAPITULO2_PATH = INPUTS_DIR / "BEN2026_Capitulo2_Oferta_Demanda_por_Fonte.xlsx"
CAPITULO3_PATH = INPUTS_DIR / "BEN2026_Capitulo3_Consumo_Energia_por_Setor.xlsx"

FONTE_BEN = "Balanço Energético Nacional 2026 (EPE, ano-base 2025)"

# (cadeia_prioritaria, source_path, block_title, match_occurrence)
TARGET_BLOCKS: tuple[tuple[str, Path, str, int], ...] = (
    ("aco", CAPITULO3_PATH, "SETOR INDUSTRIAL - FERRO-GUSA E AÇO", 0),
    ("fertilizantes", CAPITULO3_PATH, "SETOR INDUSTRIAL - QUÍMICA", 0),
    ("silicio", CAPITULO2_PATH, "ENERGIA SOLAR FOTOVOLTAICA", 0),
    ("combustiveis_transicao", CAPITULO2_PATH, "BIODIESEL", 0),
    ("combustiveis_transicao", CAPITULO2_PATH, "ÁLCOOL ETÍLICO TOTAL (*)", 0),
)


def main() -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    workbooks: dict[Path, list[tuple]] = {}
    rows: list[dict[str, object]] = []
    summary: dict[str, int] = {}

    for chain_name, source_path, block_title, occurrence in TARGET_BLOCKS:
        if source_path not in workbooks:
            wb = openpyxl.load_workbook(source_path, read_only=True, data_only=True)
            workbooks[source_path] = list(wb[wb.sheetnames[0]].iter_rows(values_only=True))
        sheet_rows = workbooks[source_path]

        block_rows = extract_block(sheet_rows, block_title, occurrence)
        for row in block_rows:
            row["cadeia_prioritaria"] = chain_name
        rows.extend(block_rows)
        summary[f"{chain_name} / {block_title}"] = len(block_rows)

    output_sql = OUTPUT_DIR / "load_analytical_energy_context_national.sql"
    write_sql(rows, output_sql)

    print(
        json.dumps(
            {**summary, "total_rows": len(rows), "output": str(output_sql)},
            ensure_ascii=False,
            indent=2,
        )
    )


def extract_block(sheet_rows: list[tuple], title: str, occurrence: int) -> list[dict[str, object]]:
    """Find the `occurrence`-th block whose title row matches `title`, read
    its self-documented unit, and unpivot every (fonte, ano) cell into one
    row per non-empty numeric value."""

    matches_seen = 0
    title_row_index = None
    for i, row in enumerate(sheet_rows):
        if row and row[0] == title:
            if matches_seen == occurrence:
                title_row_index = i
                break
            matches_seen += 1

    if title_row_index is None:
        raise ValueError(f"Bloco '{title}' (ocorrencia {occurrence}) nao encontrado na planilha.")

    unit = next(
        (cell for cell in sheet_rows[title_row_index] if isinstance(cell, str) and cell.startswith("UNIDADE:")),
        None,
    )
    if unit is None:
        raise ValueError(f"Bloco '{title}': celula 'UNIDADE: ...' nao encontrada na linha do titulo.")
    unit = unit.removeprefix("UNIDADE:").strip()

    header_row = sheet_rows[title_row_index + 1]
    years = [cell for cell in header_row[1:] if isinstance(cell, int)]
    year_offset = 1  # header_row[1] is the first year column

    setor_ben = title.replace(" (*)", "").strip()

    result: list[dict[str, object]] = []
    i = title_row_index + 2
    while i < len(sheet_rows):
        row = sheet_rows[i]
        label = row[0]
        if label is None:
            break
        if isinstance(label, str) and label.strip().upper().startswith("TABELA"):
            break
        values = row[year_offset : year_offset + len(years)]
        if any(isinstance(v, (int, float)) for v in values):
            fonte_energetica = label.strip()
            for year, value in zip(years, values):
                if isinstance(value, (int, float)):
                    result.append(
                        {
                            "ano_referencia": year,
                            "setor_ben": setor_ben,
                            "fonte_energetica": fonte_energetica,
                            "valor": float(value),
                            "unidade": unit,
                            "fonte_ben": FONTE_BEN,
                        }
                    )
        i += 1

    return result


def write_sql(rows: list[dict[str, object]], output_path: Path) -> None:
    statements = [
        "-- Gerado por build_energy_context_ben.py -- nao editar a mao.",
        "CREATE TABLE IF NOT EXISTS analytical_energy_context_national (",
        "  ano_referencia integer NOT NULL,",
        "  cadeia_prioritaria text NOT NULL,",
        "  setor_ben text NOT NULL,",
        "  fonte_energetica text NOT NULL,",
        "  valor numeric NOT NULL DEFAULT 0,",
        "  unidade text NOT NULL,",
        "  fonte_ben text NOT NULL",
        ");",
        "BEGIN;",
        # This script owns the whole table (single source, no other loader
        # writes to it), so a full truncate+reload is safe to rerun.
        "TRUNCATE analytical_energy_context_national;",
    ]

    for row in rows:
        statements.append(
            "INSERT INTO analytical_energy_context_national "
            "(ano_referencia, cadeia_prioritaria, setor_ben, fonte_energetica, valor, unidade, fonte_ben) VALUES ("
            f"{int(row['ano_referencia'])}, "
            f"'{sql_text(row['cadeia_prioritaria'])}', "
            f"'{sql_text(row['setor_ben'])}', "
            f"'{sql_text(row['fonte_energetica'])}', "
            f"{sql_number(row['valor'])}, "
            f"'{sql_text(row['unidade'])}', "
            f"'{sql_text(row['fonte_ben'])}');"
        )

    statements.append("COMMIT;")
    output_path.write_text("\n".join(statements) + "\n", encoding="utf-8")


def sql_text(value: object) -> str:
    return str(value).replace("'", "''")


def sql_number(value: object) -> str:
    return repr(float(value))


if __name__ == "__main__":
    main()
