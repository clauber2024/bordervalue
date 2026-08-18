"""Generate a small NCM/PRODLIST/CNAE code -> description lookup for the
frontend's hover tooltips (CodeList and friends).

Scoped to only the codes this app actually renders anywhere (all 4 chains'
NCM baskets, plus whatever PRODLIST/CNAE codes those NCMs bridge to) instead
of embedding the full national classification tables, which would be orders
of magnitude larger for no benefit -- nothing outside this catalog is ever
shown in the UI.

Sources, all real/official, none fabricated:
  - NCM: dados/cache/ncm_vigente.json (already used elsewhere in the pipeline)
  - PRODLIST: inputs/official/Prodlist_Industria_2022.xlsx (CONCLA structure
    file, same 2022 vintage as the existing ncm_prodlist_2022.xlsx bridge)
  - CNAE: IBGE's own classificacoes API (servicodados.ibge.gov.br/api/v2/cnae),
    fetched and cached locally once, then filtered to the classes in use

Output: data/code_descriptions.json, imported directly by the Next.js
frontend (components/CodeTooltip.tsx) -- not in inputs/official/ because
this is a derived artifact, not a raw official source file.
"""

from __future__ import annotations

import csv
import json
import re
import urllib.request
from html import unescape
from pathlib import Path

import build_solar_sovereignty_metrics as core
import build_sector_sovereignty_metrics as sector


def clean_text(value: object) -> str:
    # Same cleanup as build_matriz_auditoria_published_ncm.py's clean_text --
    # the official NCM json embeds raw HTML (e.g. "<i>spiegel</i>") in some
    # Descricao fields.
    text = unescape(str(value or ""))
    text = re.sub(r"<[^>]+>", "", text)
    return re.sub(r"\s+", " ", text).strip()

ROOT = Path(__file__).resolve().parent
NCM_VIGENTE_PATH = ROOT / "dados" / "cache" / "ncm_vigente.json"
PRODLIST_XLSX_PATH = ROOT / "inputs" / "official" / "Prodlist_Industria_2022.xlsx"
BRIDGE_PATH = ROOT / "outputs" / "official_2026" / "bridge_ncm_prodlist_cnae.csv"
CNAE_CACHE_PATH = ROOT / "dados" / "cache" / "cnae_classes_ibge.json"
OUTPUT_PATH = ROOT / "data" / "code_descriptions.json"


def collect_used_ncm_codes() -> set[str]:
    codes: set[str] = set()
    for definition in core.SOLAR_INPUTS:
        codes.update(definition.ncm_codes)
    for config in sector.CHAINS.values():
        for definition in config["definitions"]:
            codes.update(definition.ncm_codes)
    return codes


def load_ncm_descriptions(used_codes: set[str]) -> dict[str, str]:
    # NCM's own Descricao is only the incremental text at that hierarchy
    # level (e.g. the 8-digit leaf for 72083990 is literally "Outros") --
    # meaningless shown alone in a tooltip. Same hierarchy-walk already used
    # by build_matriz_auditoria_published_ncm.py's descricao_ncm_hierarquica
    # column, reused here for consistency with the rest of the pipeline.
    data = json.loads(NCM_VIGENTE_PATH.read_text(encoding="utf-8"))
    by_code: dict[str, str] = {}
    for item in data["Nomenclaturas"]:
        code = item["Codigo"].replace(".", "")
        text = clean_text(item["Descricao"])
        if code and text:
            by_code[code] = text

    result: dict[str, str] = {}
    for code in used_codes:
        if code not in by_code:
            continue
        hierarchy: list[str] = []
        for prefix_len in (2, 4, 6, 8):
            part = by_code.get(code[:prefix_len])
            if part and part not in hierarchy:
                hierarchy.append(part)
        result[code] = " > ".join(hierarchy)
    return result


def load_bridge_codes(used_ncm_codes: set[str]) -> tuple[set[str], set[str]]:
    prodlist_codes: set[str] = set()
    cnae_codes: set[str] = set()
    with BRIDGE_PATH.open("r", encoding="utf-8-sig", newline="") as handle:
        for row in csv.DictReader(handle):
            ncm = str(row.get("ncm", "")).zfill(8)
            if ncm not in used_ncm_codes:
                continue
            prodlist = row.get("prodlist_code", "")
            if prodlist:
                prodlist_codes.add(prodlist)
            cnae = str(row.get("cnae_class", ""))
            if cnae:
                cnae_codes.add(cnae)
    return prodlist_codes, cnae_codes


def load_prodlist_descriptions(used_codes: set[str]) -> dict[str, str]:
    import openpyxl

    wb = openpyxl.load_workbook(PRODLIST_XLSX_PATH, data_only=True)
    ws = wb[wb.sheetnames[0]]
    result: dict[str, str] = {}
    for row in ws.iter_rows(values_only=True):
        code = row[1] if len(row) > 1 else None
        desc = row[2] if len(row) > 2 else None
        if not code or not desc:
            continue
        code = str(code).strip()
        if code in used_codes:
            result[code] = clean_text(desc)
    return result


def fetch_cnae_classes() -> list[dict[str, object]]:
    if CNAE_CACHE_PATH.exists():
        return json.loads(CNAE_CACHE_PATH.read_text(encoding="utf-8"))
    with urllib.request.urlopen("https://servicodados.ibge.gov.br/api/v2/cnae/classes", timeout=60) as response:
        raw = json.loads(response.read().decode("utf-8"))
    CNAE_CACHE_PATH.parent.mkdir(parents=True, exist_ok=True)
    CNAE_CACHE_PATH.write_text(json.dumps(raw, ensure_ascii=False), encoding="utf-8")
    return raw


def load_cnae_descriptions(used_codes: set[str]) -> dict[str, str]:
    classes = fetch_cnae_classes()
    # IBGE's own id is "XXXX-X" (class + check digit); this app's bridge/
    # frontend only ever carries the bare 4-digit class, so key by that.
    result: dict[str, str] = {}
    for item in classes:
        raw_id = str(item.get("id", ""))
        class4 = raw_id.replace("-", "").replace(".", "")[:4]
        if class4 in used_codes and class4 not in result:
            result[class4] = clean_text(item.get("descricao", ""))
    return result


def main() -> None:
    used_ncm = collect_used_ncm_codes()
    ncm_desc = load_ncm_descriptions(used_ncm)

    used_prodlist, used_cnae = load_bridge_codes(used_ncm)
    prodlist_desc = load_prodlist_descriptions(used_prodlist)
    cnae_desc = load_cnae_descriptions(used_cnae)

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT_PATH.write_text(
        json.dumps({"ncm": ncm_desc, "prodlist": prodlist_desc, "cnae": cnae_desc}, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    print(
        json.dumps(
            {
                "ncm_used": len(used_ncm), "ncm_matched": len(ncm_desc),
                "prodlist_used": len(used_prodlist), "prodlist_matched": len(prodlist_desc),
                "cnae_used": len(used_cnae), "cnae_matched": len(cnae_desc),
                "output": str(OUTPUT_PATH),
            },
            ensure_ascii=False,
        )
    )


if __name__ == "__main__":
    main()
